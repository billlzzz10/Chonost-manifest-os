"""
🚀 Enhanced RAG System - High Performance & Deep Document Scanning.

An improved RAG system with enhanced performance and support for deep document scanning.

Features:
- Parallel Document Processing
- Intelligent File Type Detection
- Advanced Caching Strategy
- Performance Monitoring
- Deep Code Analysis
- Document Content Extraction
"""

import asyncio
import json
import logging
import hashlib
import time
from datetime import datetime
from typing import Dict, List, Optional, Any, Union, Tuple, Set
from dataclasses import dataclass, asdict
from pathlib import Path
import numpy as np
import sqlite3
import redis
import requests
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor
import threading
import multiprocessing
import mimetypes
import magic
import re
import os

# Document Processing Imports
try:
    import PyPDF2
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import docx
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from bs4 import BeautifulSoup
    HTML_AVAILABLE = True
except ImportError:
    HTML_AVAILABLE = False

# Code Analysis Imports
try:
    import ast
    import tokenize
    AST_AVAILABLE = True
except ImportError:
    AST_AVAILABLE = False

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class PerformanceMetrics:
    """
    Stores performance metrics.

    Attributes:
        start_time (float): The start time of the operation.
        end_time (float): The end time of the operation.
        files_processed (int): The number of files processed.
        documents_extracted (int): The number of documents extracted.
        embeddings_generated (int): The number of embeddings generated.
        cache_hits (int): The number of cache hits.
        cache_misses (int): The number of cache misses.
        errors (List[str]): A list of errors that occurred.
    """
    start_time: float
    end_time: float
    files_processed: int
    documents_extracted: int
    embeddings_generated: int
    cache_hits: int
    cache_misses: int
    errors: List[str]
    
    @property
    def total_time(self) -> float:
        """Calculates the total time taken for the operation."""
        return self.end_time - self.start_time
    
    @property
    def files_per_second(self) -> float:
        """Calculates the number of files processed per second."""
        return self.files_processed / self.total_time if self.total_time > 0 else 0
    
    @property
    def documents_per_second(self) -> float:
        """Calculates the number of documents extracted per second."""
        return self.documents_extracted / self.total_time if self.total_time > 0 else 0

@dataclass
class DocumentContent:
    """
    Represents the content of an extracted document.

    Attributes:
        file_path (str): The path to the file.
        content_type (str): The type of content ('text', 'code', 'document', 'image', 'binary').
        content (str): The extracted content.
        metadata (Dict[str, Any]): Metadata for the document.
        extracted_at (datetime): The timestamp when the document was extracted.
        file_size (int): The size of the file in bytes.
        processing_time (float): The time taken to process the file.
    """
    file_path: str
    content_type: str  # 'text', 'code', 'document', 'image', 'binary'
    content: str
    metadata: Dict[str, Any]
    extracted_at: datetime
    file_size: int
    processing_time: float

class IntelligentFileProcessor:
    """
    An intelligent file processing system.

    Attributes:
        config (Dict[str, Any]): The configuration for the file processor.
        supported_extensions (Dict[str, str]): A dictionary of supported file extensions and their content types.
        size_limits (Dict[str, int]): A dictionary of file size limits for different content types.
        max_workers (int): The maximum number of workers for parallel processing.
        chunk_size (int): The size of chunks for processing large files.
        chunk_overlap (int): The overlap between chunks.
    """
    
    def __init__(self, config: Dict[str, Any] = None):
        """
        Initializes the IntelligentFileProcessor.

        Args:
            config (Dict[str, Any], optional): The configuration for the file processor. Defaults to None.
        """
        self.config = config or {}
        self.supported_extensions = {
            # Text & Code Files
            '.txt': 'text', '.md': 'text', '.rst': 'text', '.log': 'text',
            '.py': 'code', '.js': 'code', '.ts': 'code', '.jsx': 'code', '.tsx': 'code',
            '.java': 'code', '.cpp': 'code', '.c': 'code', '.h': 'code', '.hpp': 'code',
            '.cs': 'code', '.php': 'code', '.rb': 'code', '.go': 'code', '.rs': 'code',
            '.swift': 'code', '.kt': 'code', '.scala': 'code', '.dart': 'code',
            '.html': 'code', '.css': 'code', '.scss': 'code', '.sass': 'code',
            '.xml': 'code', '.json': 'code', '.yaml': 'code', '.yml': 'code',
            '.toml': 'code', '.ini': 'code', '.cfg': 'code', '.conf': 'code',
            '.sh': 'code', '.bat': 'code', '.ps1': 'code', '.sql': 'code',
            
            # Document Files
            '.pdf': 'document', '.docx': 'document', '.doc': 'document',
            '.xlsx': 'document', '.xls': 'document', '.pptx': 'document',
            '.ppt': 'document', '.odt': 'document', '.ods': 'document',
            
            # Data Files
            '.csv': 'data', '.tsv': 'data', '.xml': 'data', '.json': 'data',
            
            # Configuration Files
            '.env': 'config', '.gitignore': 'config', '.dockerfile': 'config',
            '.dockerignore': 'config', 'docker-compose.yml': 'config',
            'package.json': 'config', 'requirements.txt': 'config',
            'pom.xml': 'config', 'build.gradle': 'config', 'Cargo.toml': 'config',
            'go.mod': 'config', 'composer.json': 'config', 'Gemfile': 'config'
        }
        
        # File size limits (in bytes)
        self.size_limits = {
            'text': 10 * 1024 * 1024,  # 10MB
            'code': 5 * 1024 * 1024,   # 5MB
            'document': 50 * 1024 * 1024,  # 50MB
            'data': 100 * 1024 * 1024,  # 100MB
            'config': 1 * 1024 * 1024   # 1MB
        }
        
        # Performance settings
        self.max_workers = self.config.get('max_workers', min(32, (os.cpu_count() or 1) + 4))
        self.chunk_size = self.config.get('chunk_size', 1000)
        self.chunk_overlap = self.config.get('chunk_overlap', 200)
        
        logger.info(f"🚀 Intelligent File Processor is ready (Workers: {self.max_workers})")
    
    async def process_directory_deep(self, root_path: str, 
                                   include_patterns: List[str] = None,
                                   exclude_patterns: List[str] = None) -> List[DocumentContent]:
        """
        Processes a directory deeply.

        Args:
            root_path (str): The root path of the directory to process.
            include_patterns (List[str], optional): A list of patterns to include. Defaults to None.
            exclude_patterns (List[str], optional): A list of patterns to exclude. Defaults to None.

        Returns:
            List[DocumentContent]: A list of processed document contents.
        """
        start_time = time.time()
        
        try:
            # Create patterns for filtering
            include_patterns = include_patterns or ['*']
            exclude_patterns = exclude_patterns or [
                '*.exe', '*.dll', '*.so', '*.dylib', '*.bin',
                '*.zip', '*.tar', '*.gz', '*.rar', '*.7z',
                '*.mp3', '*.mp4', '*.avi', '*.mov', '*.wmv',
                '*.jpg', '*.jpeg', '*.png', '*.gif', '*.bmp',
                '*.iso', '*.img', '*.vmdk', '*.vhd'
            ]
            
            # Find all files
            all_files = self._find_files_recursive(root_path, include_patterns, exclude_patterns)
            logger.info(f"🔍 Found {len(all_files)} files in {root_path}")
            
            # Process in parallel
            documents = await self._process_files_parallel(all_files)
            
            end_time = time.time()
            processing_time = end_time - start_time
            
            logger.info(f"✅ Processing complete: {len(documents)} documents in {processing_time:.2f} seconds")
            logger.info(f"📊 Speed: {len(documents)/processing_time:.2f} documents/second")
            
            return documents
            
        except Exception as e:
            logger.error(f"❌ Error processing directory: {e}")
            return []
    
    def _find_files_recursive(self, root_path: str, include_patterns: List[str], 
                            exclude_patterns: List[str]) -> List[str]:
        """
        Finds files recursively with pattern matching.

        Args:
            root_path (str): The root path to search in.
            include_patterns (List[str]): A list of patterns to include.
            exclude_patterns (List[str]): A list of patterns to exclude.

        Returns:
            List[str]: A list of found file paths.
        """
        files = []
        root_path = Path(root_path)
        
        try:
            for pattern in include_patterns:
                for file_path in root_path.rglob(pattern):
                    if file_path.is_file():
                        # Check exclude patterns
                        should_exclude = False
                        for exclude_pattern in exclude_patterns:
                            if file_path.match(exclude_pattern):
                                should_exclude = True
                                break
                        
                        if not should_exclude:
                            files.append(str(file_path))
            
            # Remove duplicates
            files = list(set(files))
            return files
            
        except Exception as e:
            logger.error(f"❌ Error finding files: {e}")
            return []
    
    async def _process_files_parallel(self, file_paths: List[str]) -> List[DocumentContent]:
        """
        Processes files in parallel.

        Args:
            file_paths (List[str]): A list of file paths to process.

        Returns:
            List[DocumentContent]: A list of processed document contents.
        """
        documents = []
        
        # Split files into batches
        batch_size = max(1, len(file_paths) // self.max_workers)
        batches = [file_paths[i:i + batch_size] for i in range(0, len(file_paths), batch_size)]
        
        # Process in parallel
        tasks = []
        for batch in batches:
            task = asyncio.create_task(self._process_file_batch(batch))
            tasks.append(task)
        
        # Wait for all results
        batch_results = await asyncio.gather(*tasks, return_exceptions=True)
        
        # Combine results
        for batch_result in batch_results:
            if isinstance(batch_result, list):
                documents.extend(batch_result)
            else:
                logger.error(f"❌ Batch processing error: {batch_result}")
        
        return documents
    
    async def _process_file_batch(self, file_paths: List[str]) -> List[DocumentContent]:
        """
        Processes a batch of files.

        Args:
            file_paths (List[str]): A list of file paths to process.

        Returns:
            List[DocumentContent]: A list of processed document contents.
        """
        documents = []
        
        for file_path in file_paths:
            try:
                doc = await self._process_single_file(file_path)
                if doc:
                    documents.append(doc)
            except Exception as e:
                logger.warning(f"⚠️ Could not process file {file_path}: {e}")
                continue
        
        return documents
    
    async def _process_single_file(self, file_path: str) -> Optional[DocumentContent]:
        """
        Processes a single file.

        Args:
            file_path (str): The path to the file to process.

        Returns:
            Optional[DocumentContent]: The processed document content, or None if processing fails.
        """
        start_time = time.time()
        
        try:
            file_path = Path(file_path)
            
            # Check if the file exists
            if not file_path.exists():
                return None
            
            # Check file size
            file_size = file_path.stat().st_size
            if file_size == 0:
                return None
            
            # Determine content type
            content_type = self._determine_content_type(file_path)
            if not content_type:
                return None
            
            # Check file size by type
            size_limit = self.size_limits.get(content_type, 1 * 1024 * 1024)
            if file_size > size_limit:
                logger.info(f"⚠️ Skipping file {file_path} (too large: {file_size} bytes)")
                return None
            
            # Extract content
            content = await self._extract_content(file_path, content_type)
            if not content:
                return None
            
            # Create metadata
            metadata = self._create_metadata(file_path, content_type, file_size)
            
            processing_time = time.time() - start_time
            
            return DocumentContent(
                file_path=str(file_path),
                content_type=content_type,
                content=content,
                metadata=metadata,
                extracted_at=datetime.now(),
                file_size=file_size,
                processing_time=processing_time
            )
            
        except Exception as e:
            logger.error(f"❌ Error processing file {file_path}: {e}")
            return None
    
    def _determine_content_type(self, file_path: Path) -> Optional[str]:
        """
        Determines the content type of a file.

        Args:
            file_path (Path): The path to the file.

        Returns:
            Optional[str]: The content type, or None if not determined.
        """
        file_name = file_path.name.lower()
        
        # Check by extension
        for ext, content_type in self.supported_extensions.items():
            if file_name.endswith(ext):
                return content_type
        
        # Check by file name
        for pattern, content_type in self.supported_extensions.items():
            if pattern in file_name:
                return content_type
        
        # Use magic number for unknown files
        try:
            mime_type, _ = mimetypes.guess_type(str(file_path))
            if mime_type:
                if mime_type.startswith('text/'):
                    return 'text'
                elif mime_type.startswith('application/'):
                    return 'document'
        except:
            pass
        
        return None
    
    async def _extract_content(self, file_path: Path, content_type: str) -> Optional[str]:
        """
        Extracts content from a file.

        Args:
            file_path (Path): The path to the file.
            content_type (str): The content type of the file.

        Returns:
            Optional[str]: The extracted content, or None if extraction fails.
        """
        try:
            if content_type == 'text':
                return await self._extract_text_content(file_path)
            elif content_type == 'code':
                return await self._extract_code_content(file_path)
            elif content_type == 'document':
                return await self._extract_document_content(file_path)
            elif content_type == 'data':
                return await self._extract_data_content(file_path)
            elif content_type == 'config':
                return await self._extract_config_content(file_path)
            else:
                return None
                
        except Exception as e:
            logger.error(f"❌ Error extracting content from {file_path}: {e}")
            return None
    
    async def _extract_text_content(self, file_path: Path) -> Optional[str]:
        """
        Extracts text content from a file.

        Args:
            file_path (Path): The path to the file.

        Returns:
            Optional[str]: The extracted text content, or None if extraction fails.
        """
        try:
            # Try reading as UTF-8 first
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                return content
            except UnicodeDecodeError:
                # Try other encodings
                encodings = ['latin-1', 'cp1252', 'iso-8859-1']
                for encoding in encodings:
                    try:
                        with open(file_path, 'r', encoding=encoding) as f:
                            content = f.read()
                        return content
                    except UnicodeDecodeError:
                        continue
                
                # If all else fails, read as binary and decode with ignore
                with open(file_path, 'rb') as f:
                    content = f.read().decode('utf-8', errors='ignore')
                return content
                
        except Exception as e:
            logger.error(f"❌ Could not read text file {file_path}: {e}")
            return None
    
    async def _extract_code_content(self, file_path: Path) -> Optional[str]:
        """
        Extracts code content from a file.

        Args:
            file_path (Path): The path to the file.

        Returns:
            Optional[str]: The extracted code content, or None if extraction fails.
        """
        try:
            content = await self._extract_text_content(file_path)
            if not content:
                return None
            
            # Add metadata for code
            code_metadata = self._analyze_code_structure(content, file_path.suffix)
            
            # Combine metadata with content
            enhanced_content = f"// File: {file_path.name}\n"
            enhanced_content += f"// Language: {code_metadata.get('language', 'unknown')}\n"
            enhanced_content += f"// Functions: {len(code_metadata.get('functions', []))}\n"
            enhanced_content += f"// Classes: {len(code_metadata.get('classes', []))}\n"
            enhanced_content += "// Content:\n"
            enhanced_content += content
            
            return enhanced_content
            
        except Exception as e:
            logger.error(f"❌ Could not extract code content from {file_path}: {e}")
            return None
    
    def _analyze_code_structure(self, content: str, file_extension: str) -> Dict[str, Any]:
        """
        Analyzes the structure of a code file.

        Args:
            content (str): The content of the code file.
            file_extension (str): The file extension of the code file.

        Returns:
            Dict[str, Any]: A dictionary of the code's structure.
        """
        metadata = {
            'language': self._get_language_from_extension(file_extension),
            'functions': [],
            'classes': [],
            'imports': [],
            'comments': []
        }
        
        try:
            if file_extension == '.py' and AST_AVAILABLE:
                # Analyze Python code
                try:
                    tree = ast.parse(content)
                    for node in ast.walk(tree):
                        if isinstance(node, ast.FunctionDef):
                            metadata['functions'].append(node.name)
                        elif isinstance(node, ast.ClassDef):
                            metadata['classes'].append(node.name)
                        elif isinstance(node, ast.Import):
                            for alias in node.names:
                                metadata['imports'].append(alias.name)
                        elif isinstance(node, ast.ImportFrom):
                            if node.module:
                                metadata['imports'].append(node.module)
                except:
                    pass
            
            # Find comments
            comment_patterns = {
                '.py': [r'#.*$', r'""".*?"""', r"'''.*?'''"],
                '.js': [r'//.*$', r'/\*.*?\*/'],
                '.ts': [r'//.*$', r'/\*.*?\*/'],
                '.java': [r'//.*$', r'/\*.*?\*/'],
                '.cpp': [r'//.*$', r'/\*.*?\*/'],
                '.c': [r'//.*$', r'/\*.*?\*/'],
                '.html': [r'<!--.*?-->'],
                '.css': [r'/\*.*?\*/'],
                '.sql': [r'--.*$', r'/\*.*?\*/']
            }
            
            patterns = comment_patterns.get(file_extension, [])
            for pattern in patterns:
                comments = re.findall(pattern, content, re.MULTILINE | re.DOTALL)
                metadata['comments'].extend(comments)
            
        except Exception as e:
            logger.warning(f"⚠️ Could not analyze code structure: {e}")
        
        return metadata
    
    def _get_language_from_extension(self, extension: str) -> str:
        """
        Converts a file extension to a language name.

        Args:
            extension (str): The file extension.

        Returns:
            str: The language name.
        """
        language_map = {
            '.py': 'Python', '.js': 'JavaScript', '.ts': 'TypeScript',
            '.jsx': 'React JSX', '.tsx': 'React TSX', '.java': 'Java',
            '.cpp': 'C++', '.c': 'C', '.h': 'C Header', '.hpp': 'C++ Header',
            '.cs': 'C#', '.php': 'PHP', '.rb': 'Ruby', '.go': 'Go',
            '.rs': 'Rust', '.swift': 'Swift', '.kt': 'Kotlin',
            '.scala': 'Scala', '.dart': 'Dart', '.html': 'HTML',
            '.css': 'CSS', '.scss': 'SCSS', '.sass': 'Sass',
            '.xml': 'XML', '.json': 'JSON', '.yaml': 'YAML',
            '.yml': 'YAML', '.toml': 'TOML', '.ini': 'INI',
            '.cfg': 'Config', '.conf': 'Config', '.sh': 'Shell',
            '.bat': 'Batch', '.ps1': 'PowerShell', '.sql': 'SQL'
        }
        return language_map.get(extension, 'Unknown')
    
    def _create_metadata(self, file_path: Path, content_type: str, file_size: int) -> Dict[str, Any]:
        """
        Creates metadata for a file.

        Args:
            file_path (Path): The path to the file.
            content_type (str): The content type of the file.
            file_size (int): The size of the file in bytes.

        Returns:
            Dict[str, Any]: A dictionary of metadata.
        """
        try:
            stat = file_path.stat()
            
            return {
                'file_name': file_path.name,
                'file_path': str(file_path),
                'content_type': content_type,
                'file_size': file_size,
                'file_extension': file_path.suffix.lower(),
                'created_time': datetime.fromtimestamp(stat.st_ctime).isoformat(),
                'modified_time': datetime.fromtimestamp(stat.st_mtime).isoformat(),
                'accessed_time': datetime.fromtimestamp(stat.st_atime).isoformat(),
                'is_hidden': file_path.name.startswith('.'),
                'parent_directory': str(file_path.parent),
                'depth_level': len(file_path.parts) - 1
            }
            
        except Exception as e:
            logger.error(f"❌ Could not create metadata for {file_path}: {e}")
            return {}
    
    async def _extract_document_content(self, file_path: Path) -> Optional[str]:
        """
        Extracts content from a document file.

        Args:
            file_path (Path): The path to the document file.

        Returns:
            Optional[str]: The extracted content, or None if extraction fails.
        """
        try:
            extension = file_path.suffix.lower()
            
            if extension == '.pdf' and PDF_AVAILABLE:
                return await self._extract_pdf_content(file_path)
            elif extension == '.docx' and DOCX_AVAILABLE:
                return await self._extract_docx_content(file_path)
            elif extension in ['.xlsx', '.xls'] and EXCEL_AVAILABLE:
                return await self._extract_excel_content(file_path)
            elif extension in ['.html', '.htm'] and HTML_AVAILABLE:
                return await self._extract_html_content(file_path)
            else:
                # Fallback to text extraction
                return await self._extract_text_content(file_path)
                
        except Exception as e:
            logger.error(f"❌ Could not extract document content from {file_path}: {e}")
            return None
    
    async def _extract_pdf_content(self, file_path: Path) -> Optional[str]:
        """
        Extracts content from a PDF file.

        Args:
            file_path (Path): The path to the PDF file.

        Returns:
            Optional[str]: The extracted content, or None if extraction fails.
        """
        try:
            loop = asyncio.get_event_loop()
            with ThreadPoolExecutor() as executor:
                content = await loop.run_in_executor(executor, self._extract_pdf_sync, file_path)
            return content
        except Exception as e:
            logger.error(f"❌ Could not extract PDF content from {file_path}: {e}")
            return None
    
    def _extract_pdf_sync(self, file_path: Path) -> Optional[str]:
        """
        Extracts content from a PDF file (synchronous version).

        Args:
            file_path (Path): The path to the PDF file.

        Returns:
            Optional[str]: The extracted content, or None if extraction fails.
        """
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                content = ""
                for page in pdf_reader.pages:
                    content += page.extract_text() + "\n"
                return content
        except Exception as e:
            logger.error(f"❌ PDF extraction error: {e}")
            return None
    
    async def _extract_docx_content(self, file_path: Path) -> Optional[str]:
        """
        Extracts content from a DOCX file.

        Args:
            file_path (Path): The path to the DOCX file.

        Returns:
            Optional[str]: The extracted content, or None if extraction fails.
        """
        try:
            loop = asyncio.get_event_loop()
            with ThreadPoolExecutor() as executor:
                content = await loop.run_in_executor(executor, self._extract_docx_sync, file_path)
            return content
        except Exception as e:
            logger.error(f"❌ Could not extract DOCX content from {file_path}: {e}")
            return None
    
    def _extract_docx_sync(self, file_path: Path) -> Optional[str]:
        """
        Extracts content from a DOCX file (synchronous version).

        Args:
            file_path (Path): The path to the DOCX file.

        Returns:
            Optional[str]: The extracted content, or None if extraction fails.
        """
        try:
            doc = docx.Document(file_path)
            content = ""
            for paragraph in doc.paragraphs:
                content += paragraph.text + "\n"
            return content
        except Exception as e:
            logger.error(f"❌ DOCX extraction error: {e}")
            return None
    
    async def _extract_excel_content(self, file_path: Path) -> Optional[str]:
        """
        Extracts content from an Excel file.

        Args:
            file_path (Path): The path to the Excel file.

        Returns:
            Optional[str]: The extracted content, or None if extraction fails.
        """
        try:
            loop = asyncio.get_event_loop()
            with ThreadPoolExecutor() as executor:
                content = await loop.run_in_executor(executor, self._extract_excel_sync, file_path)
            return content
        except Exception as e:
            logger.error(f"❌ Could not extract Excel content from {file_path}: {e}")
            return None
    
    def _extract_excel_sync(self, file_path: Path) -> Optional[str]:
        """
        Extracts content from an Excel file (synchronous version).

        Args:
            file_path (Path): The path to the Excel file.

        Returns:
            Optional[str]: The extracted content, or None if extraction fails.
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            content = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                content += f"Sheet: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    row_content = [str(cell) if cell is not None else "" for cell in row]
                    content += "\t".join(row_content) + "\n"
                content += "\n"
            return content
        except Exception as e:
            logger.error(f"❌ Excel extraction error: {e}")
            return None
    
    async def _extract_html_content(self, file_path: Path) -> Optional[str]:
        """
        Extracts content from an HTML file.

        Args:
            file_path (Path): The path to the HTML file.

        Returns:
            Optional[str]: The extracted content, or None if extraction fails.
        """
        try:
            content = await self._extract_text_content(file_path)
            if not content:
                return None
            
            # Use BeautifulSoup to extract text
            soup = BeautifulSoup(content, 'html.parser')
            
            # Remove script and style tags
            for script in soup(["script", "style"]):
                script.decompose()
            
            # Extract text
            text = soup.get_text()
            
            # Clean up text
            lines = (line.strip() for line in text.splitlines())
            chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
            text = ' '.join(chunk for chunk in chunks if chunk)
            
            return text
            
        except Exception as e:
            logger.error(f"❌ Could not extract HTML content from {file_path}: {e}")
            return None
    
    async def _extract_data_content(self, file_path: Path) -> Optional[str]:
        """
        Extracts content from a data file.

        Args:
            file_path (Path): The path to the data file.

        Returns:
            Optional[str]: The extracted content, or None if extraction fails.
        """
        try:
            extension = file_path.suffix.lower()
            
            if extension == '.csv':
                return await self._extract_csv_content(file_path)
            elif extension == '.json':
                return await self._extract_json_content(file_path)
            elif extension == '.xml':
                return await self._extract_xml_content(file_path)
            else:
                return await self._extract_text_content(file_path)
                
        except Exception as e:
            logger.error(f"❌ Could not extract data content from {file_path}: {e}")
            return None
    
    async def _extract_csv_content(self, file_path: Path) -> Optional[str]:
        """
        Extracts content from a CSV file.

        Args:
            file_path (Path): The path to the CSV file.

        Returns:
            Optional[str]: The extracted content, or None if extraction fails.
        """
        try:
            content = await self._extract_text_content(file_path)
            if not content:
                return None
            
            # Add metadata
            lines = content.split('\n')
            if lines:
                headers = lines[0].split(',')
                content = f"CSV File: {file_path.name}\n"
                content += f"Columns: {len(headers)}\n"
                content += f"Headers: {', '.join(headers)}\n"
                content += f"Rows: {len(lines) - 1}\n"
                content += "Content:\n" + '\n'.join(lines[:100])  # Limit to the first 100 rows
            
            return content
            
        except Exception as e:
            logger.error(f"❌ Could not extract CSV content from {file_path}: {e}")
            return None
    
    async def _extract_json_content(self, file_path: Path) -> Optional[str]:
        """
        Extracts content from a JSON file.

        Args:
            file_path (Path): The path to the JSON file.

        Returns:
            Optional[str]: The extracted content, or None if extraction fails.
        """
        try:
            content = await self._extract_text_content(file_path)
            if not content:
                return None
            
            # Parse JSON to add metadata
            try:
                data = json.loads(content)
                content = f"JSON File: {file_path.name}\n"
                content += f"Type: {type(data).__name__}\n"
                if isinstance(data, dict):
                    content += f"Keys: {', '.join(data.keys())}\n"
                elif isinstance(data, list):
                    content += f"Items: {len(data)}\n"
                content += "Content:\n" + content
            
            except json.JSONDecodeError:
                # If parsing fails, use the original content
                pass
            
            return content
            
        except Exception as e:
            logger.error(f"❌ Could not extract JSON content from {file_path}: {e}")
            return None
    
    async def _extract_xml_content(self, file_path: Path) -> Optional[str]:
        """
        Extracts content from an XML file.

        Args:
            file_path (Path): The path to the XML file.

        Returns:
            Optional[str]: The extracted content, or None if extraction fails.
        """
        try:
            content = await self._extract_text_content(file_path)
            if not content:
                return None
            
            # Use BeautifulSoup to parse XML
            soup = BeautifulSoup(content, 'xml')
            
            # Extract important information
            content = f"XML File: {file_path.name}\n"
            content += f"Root Element: {soup.find().name if soup.find() else 'None'}\n"
            content += "Content:\n" + soup.get_text()
            
            return content
            
        except Exception as e:
            logger.error(f"❌ Could not extract XML content from {file_path}: {e}")
            return None
    
    async def _extract_config_content(self, file_path: Path) -> Optional[str]:
        """
        Extracts content from a config file.

        Args:
            file_path (Path): The path to the config file.

        Returns:
            Optional[str]: The extracted content, or None if extraction fails.
        """
        try:
            content = await self._extract_text_content(file_path)
            if not content:
                return None
            
            # Add metadata for config files
            content = f"Config File: {file_path.name}\n"
            content += f"Type: Configuration\n"
            content += "Content:\n" + content
            
            return content
            
        except Exception as e:
            logger.error(f"❌ Could not extract config content from {file_path}: {e}")
            return None

class EnhancedRAGSystem:
    """
    A high-performance Enhanced RAG System.

    Attributes:
        config (Dict[str, Any]): The configuration for the RAG system.
        file_processor (IntelligentFileProcessor): The file processor instance.
        metrics (PerformanceMetrics): The performance metrics for the system.
        cache: The Redis cache instance.
    """
    
    def __init__(self, config: Dict[str, Any] = None):
        """
        Initializes the EnhancedRAGSystem.

        Args:
            config (Dict[str, Any], optional): The configuration for the RAG system. Defaults to None.
        """
        self.config = config or {}
        
        # Initialize components
        self.file_processor = IntelligentFileProcessor(config.get('file_processor', {}))
        
        # Performance monitoring
        self.metrics = PerformanceMetrics(
            start_time=0,
            end_time=0,
            files_processed=0,
            documents_extracted=0,
            embeddings_generated=0,
            cache_hits=0,
            cache_misses=0,
            errors=[]
        )
        
        # Cache system
        self.cache = redis.Redis(
            host=self.config.get("redis_host", "localhost"),
            port=self.config.get("redis_port", 6379),
            db=self.config.get("redis_db", 1),  # Use DB 1 for the enhanced system
            decode_responses=True
        )
        
        logger.info("🚀 Enhanced RAG System is ready")
    
    async def scan_directory_deep(self, root_path: str, 
                                include_patterns: List[str] = None,
                                exclude_patterns: List[str] = None,
                                max_depth: int = 10) -> Dict[str, Any]:
        """
        Scans a directory deeply.

        Args:
            root_path (str): The root path of the directory to scan.
            include_patterns (List[str], optional): A list of patterns to include. Defaults to None.
            exclude_patterns (List[str], optional): A list of patterns to exclude. Defaults to None.
            max_depth (int, optional): The maximum depth to scan. Defaults to 10.

        Returns:
            Dict[str, Any]: A report of the scan.
        """
        start_time = time.time()
        self.metrics.start_time = start_time
        
        try:
            logger.info(f"🔍 Starting deep scan of directory: {root_path}")
            
            # Process files
            documents = await self.file_processor.process_directory_deep(
                root_path, include_patterns, exclude_patterns
            )
            
            # Update metrics
            self.metrics.files_processed = len(documents)
            self.metrics.documents_extracted = len(documents)
            
            # Cache documents
            await self._cache_documents(documents)
            
            # Generate embeddings (if needed)
            if self.config.get('generate_embeddings', True):
                await self._generate_embeddings_batch(documents)
            
            end_time = time.time()
            self.metrics.end_time = end_time
            
            # Create report
            report = self._create_scan_report(documents)
            
            logger.info(f"✅ Scan complete: {len(documents)} documents in {end_time - start_time:.2f} seconds")
            
            return report
            
        except Exception as e:
            logger.error(f"❌ Error during scan: {e}")
            self.metrics.errors.append(str(e))
            return {'error': str(e)}
    
    async def _cache_documents(self, documents: List[DocumentContent]):
        """
        Caches documents.

        Args:
            documents (List[DocumentContent]): A list of documents to cache.
        """
        try:
            for doc in documents:
                cache_key = f"doc:{hashlib.md5(doc.file_path.encode()).hexdigest()}"
                cache_data = {
                    'file_path': doc.file_path,
                    'content_type': doc.content_type,
                    'content': doc.content[:10000],  # Limit size
                    'metadata': doc.metadata,
                    'extracted_at': doc.extracted_at.isoformat(),
                    'file_size': doc.file_size,
                    'processing_time': doc.processing_time
                }
                
                # Cache for 1 hour
                self.cache.setex(cache_key, 3600, json.dumps(cache_data))
            
            logger.info(f"✅ Cached {len(documents)} documents")
            
        except Exception as e:
            logger.error(f"❌ Error during caching: {e}")
    
    async def _generate_embeddings_batch(self, documents: List[DocumentContent]):
        """
        Generates embeddings in a batch.

        Args:
            documents (List[DocumentContent]): A list of documents to generate embeddings for.
        """
        try:
            # Use the original RAG system
            from .rag_system import RAGSystem
            
            rag = RAGSystem(self.config)
            
            for doc in documents:
                if doc.content:
                    success = await rag.add_document(
                        content=doc.content,
                        metadata=doc.metadata,
                        source=doc.file_path
                    )
                    if success:
                        self.metrics.embeddings_generated += 1
            
            logger.info(f"✅ Generated embeddings for {self.metrics.embeddings_generated} documents")
            
        except Exception as e:
            logger.error(f"❌ Error generating embeddings: {e}")
    
    def _create_scan_report(self, documents: List[DocumentContent]) -> Dict[str, Any]:
        """
        Creates a scan report.

        Args:
            documents (List[DocumentContent]): A list of processed documents.

        Returns:
            Dict[str, Any]: A dictionary representing the scan report.
        """
        try:
            # Group by content type
            content_types = {}
            file_extensions = {}
            total_size = 0
            
            for doc in documents:
                # Content types
                content_type = doc.content_type
                if content_type not in content_types:
                    content_types[content_type] = 0
                content_types[content_type] += 1
                
                # File extensions
                ext = doc.metadata.get('file_extension', 'unknown')
                if ext not in file_extensions:
                    file_extensions[ext] = 0
                file_extensions[ext] += 1
                
                # Total size
                total_size += doc.file_size
            
            # Performance metrics
            performance = {
                'total_time': self.metrics.total_time,
                'files_per_second': self.metrics.files_per_second,
                'documents_per_second': self.metrics.documents_per_second,
                'embeddings_generated': self.metrics.embeddings_generated,
                'cache_hits': self.metrics.cache_hits,
                'cache_misses': self.metrics.cache_misses
            }
            
            return {
                'summary': {
                    'total_documents': len(documents),
                    'total_size_bytes': total_size,
                    'total_size_mb': total_size / (1024 * 1024),
                    'content_types': content_types,
                    'file_extensions': dict(sorted(file_extensions.items(), key=lambda x: x[1], reverse=True)[:20]),
                    'performance': performance
                },
                'documents': [
                    {
                        'file_path': doc.file_path,
                        'content_type': doc.content_type,
                        'file_size': doc.file_size,
                        'processing_time': doc.processing_time,
                        'metadata': doc.metadata
                    }
                    for doc in documents[:100]  # Limit to the first 100 documents
                ],
                'errors': self.metrics.errors
            }
            
        except Exception as e:
            logger.error(f"❌ Error creating report: {e}")
            return {'error': str(e)}
    
    async def search_documents(self, query: str, content_types: List[str] = None, 
                             top_k: int = 10) -> List[Dict[str, Any]]:
        """
        Searches for documents.

        Args:
            query (str): The search query.
            content_types (List[str], optional): A list of content types to search in. Defaults to None.
            top_k (int, optional): The number of results to return. Defaults to 10.

        Returns:
            List[Dict[str, Any]]: A list of search results.
        """
        try:
            # Check cache
            cache_key = f"search:{hashlib.md5(query.encode()).hexdigest()}"
            cached_result = self.cache.get(cache_key)
            
            if cached_result:
                self.metrics.cache_hits += 1
                logger.info("✅ Using result from cache")
                return json.loads(cached_result)
            
            self.metrics.cache_misses += 1
            
            # Search in the RAG system
            from .rag_system import RAGSystem
            rag = RAGSystem(self.config)
            
            results = await rag.search(query, top_k)
            
            # Convert results
            search_results = []
            for result in results:
                search_results.append({
                    'file_path': result.document.source,
                    'content': result.document.content[:500],  # Limit size
                    'score': result.score,
                    'relevance': result.relevance,
                    'metadata': result.document.metadata
                })
            
            # Cache results
            self.cache.setex(cache_key, 1800, json.dumps(search_results))  # 30 minutes
            
            return search_results
            
        except Exception as e:
            logger.error(f"❌ Error during search: {e}")
            return []
    
    def get_performance_metrics(self) -> Dict[str, Any]:
        """
        Gets performance metrics.

        Returns:
            Dict[str, Any]: A dictionary of performance metrics.
        """
        return {
            'total_time': self.metrics.total_time,
            'files_processed': self.metrics.files_processed,
            'documents_extracted': self.metrics.documents_extracted,
            'embeddings_generated': self.metrics.embeddings_generated,
            'files_per_second': self.metrics.files_per_second,
            'documents_per_second': self.metrics.documents_per_second,
            'cache_hits': self.metrics.cache_hits,
            'cache_misses': self.metrics.cache_misses,
            'cache_hit_rate': self.metrics.cache_hits / (self.metrics.cache_hits + self.metrics.cache_misses) if (self.metrics.cache_hits + self.metrics.cache_misses) > 0 else 0,
            'errors': self.metrics.errors
        }
